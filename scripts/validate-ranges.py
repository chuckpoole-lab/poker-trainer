"""
Poker Trainer — Range Validation Script
Generates every scenario the app can produce, evaluates each decision,
and flags likely errors using independent GTO heuristics.
Outputs an Excel spreadsheet with: all decisions + flagged mismatches.
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ═══════════════════════════════════════════════════════════════
# RANGE DATA — exact copy from range-tables.ts
# ═══════════════════════════════════════════════════════════════

RANKS = ['A','K','Q','J','T','9','8','7','6','5','4','3','2']
RANK_VAL = {'A':14,'K':13,'Q':12,'J':11,'T':10,'9':9,'8':8,'7':7,'6':6,'5':5,'4':4,'3':3,'2':2}

def generate_all_hands():
    hands = []
    for i in range(len(RANKS)):
        for j in range(i, len(RANKS)):
            if i == j:
                hands.append({'code': f"{RANKS[i]}{RANKS[j]}", 'rank1': RANKS[i], 'rank2': RANKS[j], 'suited': False, 'pair': True})
            else:
                hands.append({'code': f"{RANKS[i]}{RANKS[j]}s", 'rank1': RANKS[i], 'rank2': RANKS[j], 'suited': True, 'pair': False})
                hands.append({'code': f"{RANKS[i]}{RANKS[j]}o", 'rank1': RANKS[i], 'rank2': RANKS[j], 'suited': False, 'pair': False})
    return hands

ALL_HANDS = generate_all_hands()

def hand_strength(h):
    hi = RANK_VAL[h['rank1']]
    lo = RANK_VAL[h['rank2']]
    score = hi * 15 + lo
    if h['pair']: score += 200
    if h['suited']: score += 30
    return score

def parse_range(notation):
    result = set()
    parts = [s.strip() for s in notation.split(',') if s.strip()]
    for part in parts:
        if '-' in part:
            low, high = part.split('-')
            if len(low) == 2 and low[0] == low[1] and len(high) == 2 and high[0] == high[1]:
                min_val = RANK_VAL[low[0]]
                max_val = RANK_VAL[high[0]]
                for r in RANKS:
                    if min_val <= RANK_VAL[r] <= max_val:
                        result.add(f"{r}{r}")
            elif len(low) == 3 and len(high) == 3:
                r1 = low[0]
                typ = low[2]
                min_r2 = RANK_VAL[low[1]]
                max_r2 = RANK_VAL[high[1]]
                for r in RANKS:
                    if min_r2 <= RANK_VAL[r] <= max_r2:
                        result.add(f"{r1}{r}{typ}")
        elif part.endswith('+'):
            base = part[:-1]
            if len(base) == 2 and base[0] == base[1]:
                min_val = RANK_VAL[base[0]]
                for r in RANKS:
                    if RANK_VAL[r] >= min_val:
                        result.add(f"{r}{r}")
            elif len(base) == 3:
                r1, r2, typ = base[0], base[1], base[2]
                r1_val = RANK_VAL[r1]
                min_r2 = RANK_VAL[r2]
                for r in RANKS:
                    if min_r2 <= RANK_VAL[r] < r1_val:
                        result.add(f"{r1}{r}{typ}")
        elif part in ('Ax','Kx','Qx','Jx'):
            hi = part[0]
            for r in RANKS:
                if r != hi and RANK_VAL[r] < RANK_VAL[hi]:
                    result.add(f"{hi}{r}s")
                    result.add(f"{hi}{r}o")
        else:
            result.add(part)
    return result

# Opening ranges
OPENING_RAW = {
    10: {
        'utg': {'open': '', 'jam': '22+, A2s+, A7o+, K9s+, KTo+, QTs+, JTs'},
        'hj':  {'open': '', 'jam': '22+, A2s+, A2o+, K4s+, K9o+, Q8s+, QTo+, J8s+, JTo, T8s+, 98s'},
        'co':  {'open': '', 'jam': '22+, A2s+, A2o+, K2s+, K5o+, Q5s+, Q9o+, J7s+, J9o+, T7s+, T9o, 97s+, 87s, 76s'},
        'btn': {'open': '', 'jam': '22+, A2s+, A2o+, K2s+, K2o+, Q2s+, Q5o+, J4s+, J8o+, T6s+, T8o+, 96s+, 86s+, 76s, 65s'},
        'sb':  {'open': '', 'jam': '22+, A2s+, A2o+, K2s+, K2o+, Q2s+, Q3o+, J2s+, J5o+, T4s+, T7o+, 95s+, 97o+, 85s+, 87o, 74s+, 64s+, 54s'},
    },
    15: {
        'utg': {'open': '', 'jam': '22+, A2s+, A9o+, K9s+, KTo+, QTs+, JTs'},
        'hj':  {'open': '99+, AJs+, AQo+, KQs', 'jam': '22+, A2s+, A5o+, K8s+, KTo+, Q9s+, QJo, J9s+, T9s'},
        'co':  {'open': 'TT+, ATs+, AJo+, KTs+, KQo, QJs', 'jam': '22+, A2s+, A2o+, K5s+, K9o+, Q8s+, QTo+, J8s+, JTo, T8s+, 98s, 87s'},
        'btn': {'open': 'TT+, A9s+, ATo+, KTs+, KJo+, QTs+, JTs', 'jam': '22+, A2s+, A2o+, K2s+, K5o+, Q6s+, Q9o+, J7s+, J9o+, T7s+, 97s+, 87s, 76s'},
        'sb':  {'open': '', 'jam': '22+, A2s+, A2o+, K2s+, K3o+, Q2s+, Q7o+, J4s+, J8o+, T6s+, T8o+, 96s+, 85s+, 75s+, 64s+, 54s'},
    },
    20: {
        'utg': {'open': '66+, A3s+, K8s+, Q9s+, J9s+, T9s, ATo+, KTo+, QJo', 'jam': ''},
        'hj':  {'open': '44+, A2s+, K6s+, Q8s+, J8s+, T8s+, 98s, A8o+, KTo+, QTo+, JTo', 'jam': '22, 33, A5o, A4o, A3o, A2o'},
        'co':  {'open': '22+, A2s+, K3s+, Q6s+, J7s+, T7s+, 97s+, 87s, 76s, A5o+, K9o+, QTo+, JTo', 'jam': ''},
        'btn': {'open': '22+, A2s+, K2s+, Q3s+, J5s+, T6s+, 96s+, 86s+, 75s+, 65s, 54s, A2o+, K7o+, Q9o+, J9o+, T8o+, 98o', 'jam': ''},
        'sb':  {'open': '22+, A2s+, K2s+, Q4s+, J7s+, T7s+, 97s+, 86s+, 76s, 65s, A2o+, K5o+, Q8o+, J9o+, T9o', 'jam': ''},
    },
    25: {
        'utg': {'open': '55+, A3s+, K8s+, Q9s+, J9s+, T9s, ATo+, KJo+', 'jam': ''},
        'hj':  {'open': '44+, A2s+, K6s+, Q8s+, J8s+, T8s+, 98s, A9o+, KTo+, QTo+, JTo', 'jam': ''},
        'co':  {'open': '22+, A2s+, K4s+, Q7s+, J7s+, T7s+, 97s+, 87s, 76s, A7o+, K9o+, QTo+, JTo', 'jam': ''},
        'btn': {'open': '22+, A2s+, K2s+, Q2s+, J4s+, T5s+, 95s+, 85s+, 75s+, 64s+, 54s, 43s, A2o+, K5o+, Q8o+, J8o+, T8o+, 97o+, 87o', 'jam': ''},
        'sb':  {'open': '22+, A2s+, K2s+, Q3s+, J5s+, T6s+, 96s+, 86s+, 75s+, 65s, 54s, A2o+, K4o+, Q7o+, J8o+, T8o+, 98o', 'jam': ''},
    },
    30: {
        'utg': {'open': '66+, A2s+, K9s+, QTs+, JTs, T9s, ATo+, KJo+, QJo', 'jam': ''},
        'hj':  {'open': '44+, A2s+, K7s+, Q9s+, J9s+, T8s+, 98s, 87s, A9o+, KTo+, QTo+, JTo', 'jam': ''},
        'co':  {'open': '22+, A2s+, K3s+, Q6s+, J7s+, T7s+, 97s+, 86s+, 76s, 65s, A5o+, K9o+, Q9o+, J9o+, T9o', 'jam': ''},
        'btn': {'open': '22+, A2s+, K2s+, Q2s+, J3s+, T4s+, 95s+, 85s+, 74s+, 64s+, 53s+, 43s, A2o+, K5o+, Q8o+, J8o+, T7o+, 97o+, 87o, 76o', 'jam': ''},
        'sb':  {'open': '22+, A2s+, K2s+, Q2s+, J4s+, T5s+, 95s+, 85s+, 74s+, 64s+, 53s+, 43s, A2o+, K3o+, Q6o+, J8o+, T8o+, 97o+, 87o', 'jam': ''},
    },
}

FACING_OPEN_RAW = {
    # UTG opens
    'utg_hj_15': {'jam': '99+, AQs+, AQo+', 'call': ''},
    'utg_hj_20': {'jam': 'TT+, AQs+, AQo+', 'call': '99, AJs, KQs'},
    'utg_hj_25': {'jam': 'QQ+, AKs, AKo', 'call': '77-JJ, ATs+, AJo+, KQs, QJs, JTs'},
    'utg_hj_30': {'jam': 'QQ+, AKs', 'call': '55-JJ, ATs+, AJo+, KJs+, KQo, QJs, JTs, T9s'},
    'utg_co_15': {'jam': '99+, AQs+, AQo+', 'call': ''},
    'utg_co_20': {'jam': 'TT+, AJs+, AQo+', 'call': '77-99, ATs, KQs'},
    'utg_co_25': {'jam': 'TT+, AJs+, AQo+, KQs', 'call': '66-99, ATs, AJo, KJs+, KQo, QJs, JTs, T9s'},
    'utg_co_30': {'jam': 'TT+, AQs+, AKo', 'call': '55-99, ATs+, AJo+, KJs+, KQo, QJs, JTs, T9s'},
    'utg_btn_15': {'jam': '77+, A9s+, ATo+, KTs+, KQo, QJs', 'call': ''},
    'utg_btn_20': {'jam': 'TT+, AJs+, AQo+', 'call': '66-99, ATs, KQs, QJs'},
    'utg_btn_25': {'jam': 'QQ+, AKs, AKo', 'call': '44-JJ, ATs+, AJo+, KJs+, KQo, QTs+, JTs, T9s'},
    'utg_btn_30': {'jam': 'QQ+, AKs', 'call': '33-JJ, A9s+, ATo+, KTs+, KJo+, QTs+, JTs, T9s, 98s'},
    'utg_sb_15': {'jam': 'TT+, AQs+, AQo+', 'call': ''},
    'utg_sb_20': {'jam': 'TT+, AQs+, AQo+', 'call': ''},
    'utg_sb_25': {'jam': 'QQ+, AKs, AKo', 'call': '88-JJ, AJs+, AQo, KQs'},
    'utg_sb_30': {'jam': 'QQ+, AKs', 'call': '77-JJ, ATs+, AJo+, KQs'},
    'utg_bb_15': {'jam': '99+, AQs+, AQo+', 'call': ''},
    'utg_bb_20': {'jam': 'TT+, AQs+, AQo+', 'call': '22-99, ATs+, AJo, KJs+, KQo, QJs, JTs'},
    'utg_bb_25': {'jam': 'QQ+, AKs, AKo', 'call': '22-JJ, A8s+, ATo+, K9s+, KTo+, Q9s+, J9s+, T9s, 98s, 87s'},
    'utg_bb_30': {'jam': 'QQ+, AKs', 'call': '22-JJ, A4s+, A8o+, K7s+, KTo+, Q8s+, QTo+, J8s+, JTo, T8s+, 97s+, 87s, 76s, 65s'},
    # HJ opens
    'hj_co_15': {'jam': '88+, ATs+, AJo+, KQs, A5s', 'call': ''},
    'hj_co_20': {'jam': 'TT+, AJs+, AQo+, KQs', 'call': '77-99, ATs, KJs'},
    'hj_co_25': {'jam': 'QQ+, AKs, AKo', 'call': '55-JJ, ATs+, AJo+, KJs+, KQo, QJs, JTs, T9s'},
    'hj_co_30': {'jam': 'QQ+, AKs', 'call': '44-JJ, A9s+, ATo+, KTs+, KJo+, QTs+, JTs, T9s, 98s'},
    'hj_btn_15': {'jam': '55+, A2s+, A7o+, K9s+, KTo+, QTs+, JTs', 'call': ''},
    'hj_btn_20': {'jam': '88+, A9s+, ATo+, KQs, A5s', 'call': '22-77, KJs, QJs, JTs'},
    'hj_btn_25': {'jam': '99+, ATs+, AJo+, KQs, A5s', 'call': '22-88, A2s-A9s, ATo, K9s+, KJo+, QTs+, JTs, T9s, 98s'},
    'hj_btn_30': {'jam': 'TT+, AQs+, AKo', 'call': '22-99, A2s-AJs, ATo+, K9s+, KJo+, QTs+, QJo, JTs, T9s, 98s, 87s'},
    'hj_sb_15': {'jam': '99+, ATs+, AQo+, KQs', 'call': ''},
    'hj_sb_20': {'jam': 'TT+, AQs+, AKo', 'call': ''},
    'hj_sb_25': {'jam': 'QQ+, AKs, AKo', 'call': '77-JJ, AJs+, AQo, KQs'},
    'hj_sb_30': {'jam': 'QQ+, AKs', 'call': '66-JJ, ATs+, AJo+, KJs+, KQo, QJs'},
    'hj_bb_15': {'jam': '77+, ATs+, AJo+, KQs', 'call': ''},
    'hj_bb_20': {'jam': '99+, ATs+, AJo+, KQs', 'call': '22-88, A2s-A9s, KTs+, KQo, QTs+, JTs, T9s'},
    'hj_bb_25': {'jam': 'TT+, AQs+, AKo, A5s', 'call': '22-99, A2s-AJs, A9o-AJo, K8s+, KTo+, Q9s+, J9s+, T8s+, 98s, 87s'},
    'hj_bb_30': {'jam': 'QQ+, AKs', 'call': '22-JJ, A2s+, A7o+, K6s+, K9o+, Q8s+, QTo+, J8s+, JTo, T7s+, 97s+, 86s+, 76s, 65s'},
    # CO opens
    'co_btn_15': {'jam': '55+, A2s+, A7o+, KTs+, KJo+, QTs+', 'call': ''},
    'co_btn_20': {'jam': '88+, ATs+, AJo+, KQs, A5s', 'call': '22-77, A8s-A9s, KTs+, QJs, JTs'},
    'co_btn_25': {'jam': '99+, ATs+, AJo+, KQs, A5s, A4s', 'call': '22-88, A2s-A9s, ATo, K9s+, KJo+, QTs+, QJo, JTs, T9s, 98s, 87s'},
    'co_btn_30': {'jam': 'TT+, AQs+, AKo', 'call': '22-99, A2s+, ATo+, K8s+, KTo+, Q9s+, QJo, J9s+, T9s, 98s, 87s'},
    'co_sb_15': {'jam': '66+, A2s+, A7o+, KTs+, KJo+, QTs+', 'call': ''},
    'co_sb_20': {'jam': '99+, ATs+, AJo+, KQs, A5s', 'call': ''},
    'co_sb_25': {'jam': 'TT+, AQs+, AKo, A5s, A4s', 'call': '55-99, ATs+, AJo, KJs+, KQo, QJs'},
    'co_sb_30': {'jam': 'TT+, AQs+, AKo', 'call': '44-99, A9s+, ATo+, KTs+, KJo+, QTs+, JTs, T9s'},
    'co_bb_15': {'jam': '55+, A2s+, A5o+, K9s+, KTo+, QTs+, JTs', 'call': ''},
    'co_bb_20': {'jam': '99+, ATs+, AJo+, KQs, A5s', 'call': '22-88, A2s-A9s, K8s+, KTo+, Q9s+, J9s+, T8s+, 98s, 87s'},
    'co_bb_25': {'jam': 'TT+, AQs+, AKo, A5s', 'call': '22-99, A2s-AJs, A8o-AJo, K5s+, K9o+, Q7s+, QTo+, J7s+, JTo, T7s+, 97s+, 86s+, 76s, 65s'},
    'co_bb_30': {'jam': 'TT+, AQs+, AKo', 'call': '22-99, A2s+, A2o-AJo, K2s+, K7o+, Q5s+, Q9o+, J6s+, J9o+, T6s+, T9o, 96s+, 85s+, 75s+, 65s, 54s'},
    # BTN opens
    'btn_sb_15': {'jam': '44+, A2s+, A5o+, K9s+, KTo+, QTs+, JTs', 'call': ''},
    'btn_sb_20': {'jam': '88+, A9s+, ATo+, KJs+, KQo, A5s', 'call': ''},
    'btn_sb_25': {'jam': 'TT+, ATs+, AJo+, KQs, A5s, A4s', 'call': '22-99, A2s-A9s, ATo, K8s+, KTo+, Q9s+, J9s+, T9s, 98s'},
    'btn_sb_30': {'jam': 'TT+, AQs+, AKo, A5s', 'call': '33-99, A2s-A9s, ATo, K9s+, KTo+, Q9s+, J9s+, T9s, 98s, 87s'},
    'btn_bb_15': {'jam': '44+, A2s+, A5o+, K9s+, KTo+, QTs+, JTs', 'call': ''},
    'btn_bb_20': {'jam': '99+, ATs+, AJo+, KQs', 'call': '22-88, A2s-A9s, K5s+, KTo+, Q8s+, J8s+, T8s+, 98s, 87s'},
    'btn_bb_25': {'jam': 'TT+, AJs+, AQo+, A5s', 'call': '22-99, A2s-ATs, A2o-AJo, K2s+, K9o+, Q5s+, QTo+, J7s+, JTo, T7s+, 97s+, 87s, 76s, 65s'},
    'btn_bb_30': {'jam': 'TT+, AQs+, AKo', 'call': '22-99, A2s+, A2o-AJo, K2s+, K7o+, Q4s+, Q9o+, J6s+, J9o+, T6s+, T9o, 96s+, 86s+, 75s+, 65s, 54s'},
    # SB opens
    'sb_bb_15': {'jam': '44+, A2s+, A5o+, K8s+, KTo+, Q9s+, J9s+, T9s', 'call': ''},
    'sb_bb_20': {'jam': '77+, A2s+, A7o+, K9s+, KTo+, QTs+, JTs', 'call': '22-66, A2o-A6o, K2s-K8s, K9o, Q8s, J8s+, T8s+, 98s'},
    'sb_bb_25': {'jam': '99+, ATs+, AJo+, KQs, A5s, A4s', 'call': '22-88, A2s-A9s, A2o-ATo, K2s+, K5o+, Q5s+, Q9o+, J7s+, J9o+, T7s+, T9o, 97s+, 86s+, 76s, 65s'},
    'sb_bb_30': {'jam': 'TT+, AQs+, AKo', 'call': '22-99, A2s+, A2o-AJo, K2s+, K3o+, Q3s+, Q7o+, J5s+, J8o+, T6s+, T8o+, 96s+, 85s+, 75s+, 64s+, 54s, 43s'},
    # MP opens (legacy)
    'mp_co_15': {'jam': '99+, AQs+, AKo', 'call': ''},
    'mp_co_20': {'jam': 'TT+, AJs+, AQo+', 'call': '77-99, ATs, KQs'},
    'mp_co_25': {'jam': 'TT+, AQs+, AKo', 'call': '55-99, ATs+, AJo, KJs+, KQo, QJs, JTs, T9s'},
    'mp_co_30': {'jam': 'QQ+, AKs', 'call': '44-JJ, A9s+, ATo+, KTs+, KJo+, QTs+, JTs, T9s, 98s'},
}

# ═══════════════════════════════════════════════════════════════
# COMPILE RANGES (same logic as TypeScript)
# ═══════════════════════════════════════════════════════════════

compiled_opening = {}
for stack, positions in OPENING_RAW.items():
    for pos, ranges in positions.items():
        key = f"{pos}_{stack}"
        compiled_opening[key] = {
            'open': parse_range(ranges['open']),
            'jam': parse_range(ranges['jam']),
        }

compiled_facing = {}
for key, entry in FACING_OPEN_RAW.items():
    compiled_facing[key] = {
        'jam': parse_range(entry['jam']),
        'call': parse_range(entry['call']),
    }

def get_opening_action(position, stack_bb, hand_code):
    stacks = [10, 15, 20, 25, 30]
    closest = min(stacks, key=lambda s: abs(s - stack_bb))
    key = f"{position}_{closest}"
    ranges = compiled_opening.get(key)
    if not ranges:
        return 'fold'
    if hand_code in ranges['jam']:
        return 'jam'
    if hand_code in ranges['open']:
        return 'open'
    return 'fold'

def get_facing_open_action(opener, hero, stack_bb, hand_code):
    stacks = [15, 20, 25, 30]
    closest = min(stacks, key=lambda s: abs(s - stack_bb))
    key = f"{opener}_{hero}_{closest}"
    ranges = compiled_facing.get(key)
    if not ranges:
        return 'NO_DATA'
    if hand_code in ranges['jam']:
        return 'jam'
    if hand_code in ranges['call']:
        return 'call'
    return 'fold'

# ═══════════════════════════════════════════════════════════════
# GTO VALIDATION HEURISTICS
# Independent poker knowledge to flag likely errors
# ═══════════════════════════════════════════════════════════════

PREMIUM_HANDS = {'AA', 'KK', 'QQ', 'JJ', 'AKs', 'AKo'}
STRONG_HANDS = {'TT', '99', 'AQs', 'AQo', 'AJs', 'KQs'}
MEDIUM_PAIRS = {'88', '77', '66', '55', '44', '33', '22'}
SUITED_ACES = {f"A{r}s" for r in RANKS[1:]}
SUITED_BROADWAYS = {'KQs', 'KJs', 'KTs', 'QJs', 'QTs', 'JTs'}

def validate_opening(position, stack_bb, hand_code, app_action):
    """Flag likely errors in opening decisions."""
    flags = []
    h_str = hand_strength_simple(hand_code)

    # Premium hands should NEVER fold from any position
    if hand_code in PREMIUM_HANDS and app_action == 'fold':
        flags.append(f"CRITICAL: {hand_code} should never fold preflop from any position")

    # Strong hands should not fold from most positions
    if hand_code in STRONG_HANDS and app_action == 'fold':
        if position not in ('utg',):  # maybe tight from UTG at short stacks
            flags.append(f"ERROR: {hand_code} is too strong to fold from {position.upper()}")

    # At 10bb, most playable hands should be jam not open
    if stack_bb == 10 and app_action == 'open':
        flags.append(f"WARNING: At 10bb, should be jam-or-fold, not open")

    # At 25-30bb from BTN/CO, very wide ranges expected
    if stack_bb >= 25 and position in ('btn', 'co'):
        if h_str >= 'medium' and app_action == 'fold':
            if hand_code in MEDIUM_PAIRS or hand_code in SUITED_ACES:
                flags.append(f"WARNING: {hand_code} is likely an open from {position.upper()} at {stack_bb}bb")

    return flags

def validate_facing_open(opener, hero, stack_bb, hand_code, app_action):
    """Flag likely errors in facing-open decisions."""
    flags = []

    if app_action == 'NO_DATA':
        flags.append(f"GAP: No range data for {opener.upper()} opens, {hero.upper()} defends at {stack_bb}bb")
        # Provide expected action based on heuristics
        expected = heuristic_facing_open(opener, hero, stack_bb, hand_code)
        flags.append(f"EXPECTED: {expected} (heuristic)")
        return flags

    # Premium hands should never fold facing a single raise
    if hand_code in PREMIUM_HANDS and app_action == 'fold':
        flags.append(f"CRITICAL: {hand_code} should never fold facing a single open")

    # Strong hands should not fold facing most opens
    if hand_code in STRONG_HANDS and app_action == 'fold':
        flags.append(f"ERROR: {hand_code} is too strong to fold vs {opener.upper()} open")

    # BB defense: pairs should almost always defend vs single raiser with 15+bb
    if hero == 'bb' and stack_bb >= 15:
        if hand_code in MEDIUM_PAIRS and app_action == 'fold':
            flags.append(f"ERROR: {hand_code} in BB should defend vs {opener.upper()} open at {stack_bb}bb")
        if hand_code in SUITED_ACES and app_action == 'fold':
            flags.append(f"ERROR: Suited ace {hand_code} in BB should defend vs {opener.upper()} open at {stack_bb}bb")

    # SB defense: tighter but still should defend pairs and strong hands
    if hero == 'sb' and stack_bb >= 20:
        if hand_code in MEDIUM_PAIRS and app_action == 'fold':
            if opener in ('btn', 'co'):
                flags.append(f"WARNING: {hand_code} in SB likely defends vs {opener.upper()} at {stack_bb}bb")
        if hand_code in SUITED_BROADWAYS and app_action == 'fold':
            flags.append(f"WARNING: {hand_code} in SB should defend vs {opener.upper()} at {stack_bb}bb")

    # BTN defense vs early position: should still play strong hands
    if hero == 'btn':
        if hand_code in SUITED_BROADWAYS and app_action == 'fold':
            flags.append(f"WARNING: {hand_code} on BTN should likely call vs {opener.upper()} open")

    return flags

def heuristic_facing_open(opener, hero, stack_bb, hand_code):
    """Best-guess action when no range data exists."""
    if hand_code in PREMIUM_HANDS:
        return 'jam' if stack_bb <= 20 else 'raise'
    if hand_code in STRONG_HANDS:
        return 'jam' if stack_bb <= 15 else 'call'
    if hand_code in MEDIUM_PAIRS:
        if hero == 'bb':
            return 'call' if stack_bb >= 15 else 'fold'
        if hero == 'sb':
            return 'call' if stack_bb >= 20 and opener in ('btn','co') else 'fold'
        return 'call' if stack_bb >= 20 else 'fold'
    if hand_code in SUITED_ACES:
        if hero in ('bb', 'btn'):
            return 'call'
        return 'call' if opener in ('co','btn') else 'fold'
    if hand_code in SUITED_BROADWAYS:
        return 'call'
    return 'fold'

def hand_strength_simple(hand_code):
    if hand_code in PREMIUM_HANDS: return 'premium'
    if hand_code in STRONG_HANDS: return 'strong'
    if hand_code in MEDIUM_PAIRS: return 'medium'
    if hand_code in SUITED_ACES: return 'medium'
    if hand_code in SUITED_BROADWAYS: return 'medium'
    return 'marginal'

# ═══════════════════════════════════════════════════════════════
# GENERATE ALL SCENARIOS AND VALIDATE
# ═══════════════════════════════════════════════════════════════

POSITIONS = ['utg', 'hj', 'co', 'btn', 'sb']
STACKS = [10, 15, 20, 25, 30]

# All facing-open combos that SHOULD exist
EXPECTED_FACING_COMBOS = []
openers = ['utg', 'hj', 'co', 'btn', 'sb']
for o in openers:
    for h in ['hj', 'co', 'btn', 'sb', 'bb']:
        if h != o:  # hero can't be same as opener
            seat_order = ['utg', 'utg1', 'mp', 'lj', 'hj', 'co', 'btn', 'sb', 'bb']
            o_idx = seat_order.index(o) if o in seat_order else -1
            h_idx = seat_order.index(h) if h in seat_order else -1
            if h_idx > o_idx:  # hero acts after opener
                for s in [15, 20, 25, 30]:
                    EXPECTED_FACING_COMBOS.append((o, h, s))

print(f"Validating {len(ALL_HANDS)} hands × {len(POSITIONS)} positions × {len(STACKS)} stacks (opening)")
print(f"Validating {len(ALL_HANDS)} hands × {len(EXPECTED_FACING_COMBOS)} facing-open combos")
print(f"Total scenarios: {len(ALL_HANDS) * len(POSITIONS) * len(STACKS) + len(ALL_HANDS) * len(EXPECTED_FACING_COMBOS)}")

opening_results = []
facing_results = []
all_flags = []

# 1. Validate all opening scenarios
for pos in POSITIONS:
    for stack in STACKS:
        for hand in ALL_HANDS:
            action = get_opening_action(pos, stack, hand['code'])
            flags = validate_opening(pos, stack, hand['code'], action)
            row = {
                'type': 'Opening',
                'position': pos.upper(),
                'stack': stack,
                'hand': hand['code'],
                'app_action': action.upper(),
                'flags': ' | '.join(flags) if flags else '',
                'severity': 'CRITICAL' if any('CRITICAL' in f for f in flags) else 'ERROR' if any('ERROR' in f for f in flags) else 'WARNING' if any('WARNING' in f for f in flags) else '',
            }
            opening_results.append(row)
            if flags:
                all_flags.append(row)

# 2. Validate all facing-open scenarios (including gaps)
for opener, hero, stack in EXPECTED_FACING_COMBOS:
    for hand in ALL_HANDS:
        action = get_facing_open_action(opener, hero, stack, hand['code'])
        flags = validate_facing_open(opener, hero, stack, hand['code'], action)
        row = {
            'type': 'Facing Open',
            'opener': opener.upper(),
            'position': hero.upper(),
            'stack': stack,
            'hand': hand['code'],
            'app_action': action.upper() if action != 'NO_DATA' else 'NO DATA',
            'flags': ' | '.join(flags) if flags else '',
            'severity': 'GAP' if action == 'NO_DATA' else 'CRITICAL' if any('CRITICAL' in f for f in flags) else 'ERROR' if any('ERROR' in f for f in flags) else 'WARNING' if any('WARNING' in f for f in flags) else '',
        }
        facing_results.append(row)
        if flags:
            all_flags.append(row)

print(f"\nResults:")
print(f"  Opening scenarios: {len(opening_results)}")
print(f"  Facing-open scenarios: {len(facing_results)}")
print(f"  Total flagged issues: {len(all_flags)}")

critical = sum(1 for f in all_flags if f['severity'] == 'CRITICAL')
errors = sum(1 for f in all_flags if f['severity'] == 'ERROR')
warnings = sum(1 for f in all_flags if f['severity'] == 'WARNING')
gaps = sum(1 for f in all_flags if f['severity'] == 'GAP')

print(f"  CRITICAL: {critical}")
print(f"  ERROR: {errors}")
print(f"  WARNING: {warnings}")
print(f"  GAP (missing data): {gaps}")

# Count unique gap combos
gap_combos = set()
for f in all_flags:
    if f['severity'] == 'GAP':
        gap_combos.add(f"{f.get('opener','')}->{f['position']} at {f['stack']}bb")

print(f"\n  Missing facing-open combos: {len(gap_combos)} of {len(EXPECTED_FACING_COMBOS)}")
for g in sorted(gap_combos):
    print(f"    - {g}")

# ═══════════════════════════════════════════════════════════════
# OUTPUT TO EXCEL
# ═══════════════════════════════════════════════════════════════

wb = Workbook()

# --- Sheet 1: Summary ---
ws_summary = wb.active
ws_summary.title = "Summary"

header_font = Font(name='Arial', bold=True, size=12)
title_font = Font(name='Arial', bold=True, size=14)
normal_font = Font(name='Arial', size=11)
critical_fill = PatternFill('solid', fgColor='FF4444')
error_fill = PatternFill('solid', fgColor='FFAA44')
warning_fill = PatternFill('solid', fgColor='FFEE44')
gap_fill = PatternFill('solid', fgColor='CC99FF')
white_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)

ws_summary['A1'] = 'Poker Trainer — Range Validation Report'
ws_summary['A1'].font = title_font
ws_summary['A2'] = f'Generated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'
ws_summary['A2'].font = normal_font

ws_summary['A4'] = 'Metric'
ws_summary['B4'] = 'Count'
ws_summary['A4'].font = header_font
ws_summary['B4'].font = header_font

summary_data = [
    ('Total opening scenarios tested', len(opening_results)),
    ('Total facing-open scenarios tested', len(facing_results)),
    ('', ''),
    ('CRITICAL issues (app gives clearly wrong answer)', critical),
    ('ERROR issues (likely wrong based on GTO)', errors),
    ('WARNING issues (possibly suboptimal)', warnings),
    ('GAP issues (no range data — app defaults to fold)', gaps),
    ('', ''),
    ('Missing facing-open position combos', len(gap_combos)),
    ('Existing facing-open combos', len(FACING_OPEN_RAW)),
    ('Total facing-open combos needed', len(EXPECTED_FACING_COMBOS)),
]

for i, (label, value) in enumerate(summary_data, 5):
    ws_summary[f'A{i}'] = label
    ws_summary[f'B{i}'] = value
    ws_summary[f'A{i}'].font = normal_font
    ws_summary[f'B{i}'].font = Font(name='Arial', bold=True, size=11)
    if 'CRITICAL' in str(label):
        ws_summary[f'B{i}'].fill = critical_fill
        ws_summary[f'B{i}'].font = white_font
    elif 'ERROR' in str(label):
        ws_summary[f'B{i}'].fill = error_fill
    elif 'WARNING' in str(label):
        ws_summary[f'B{i}'].fill = warning_fill
    elif 'GAP' in str(label):
        ws_summary[f'B{i}'].fill = gap_fill

ws_summary.column_dimensions['A'].width = 52
ws_summary.column_dimensions['B'].width = 14

# List missing combos
row = len(summary_data) + 7
ws_summary[f'A{row}'] = 'Missing Facing-Open Combos (app has no data for these)'
ws_summary[f'A{row}'].font = header_font
row += 1
for g in sorted(gap_combos):
    ws_summary[f'A{row}'] = g
    ws_summary[f'A{row}'].font = normal_font
    ws_summary[f'A{row}'].fill = gap_fill
    row += 1

# --- Sheet 2: Flagged Issues (only problems) ---
ws_flags = wb.create_sheet("Flagged Issues")

flag_headers = ['Severity', 'Type', 'Opener', 'Hero Position', 'Stack (BB)', 'Hand', 'App Says', 'Issue']
for col, h in enumerate(flag_headers, 1):
    cell = ws_flags.cell(row=1, column=col, value=h)
    cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='333333')
    cell.alignment = Alignment(horizontal='center')

# Sort: CRITICAL first, then ERROR, WARNING, GAP
severity_order = {'CRITICAL': 0, 'ERROR': 1, 'WARNING': 2, 'GAP': 3}
sorted_flags = sorted(all_flags, key=lambda f: (severity_order.get(f['severity'], 99), f['hand']))

for i, f in enumerate(sorted_flags, 2):
    ws_flags.cell(row=i, column=1, value=f['severity']).font = normal_font
    ws_flags.cell(row=i, column=2, value=f['type']).font = normal_font
    ws_flags.cell(row=i, column=3, value=f.get('opener', '')).font = normal_font
    ws_flags.cell(row=i, column=4, value=f['position']).font = normal_font
    ws_flags.cell(row=i, column=5, value=f['stack']).font = normal_font
    ws_flags.cell(row=i, column=6, value=f['hand']).font = normal_font
    ws_flags.cell(row=i, column=7, value=f['app_action']).font = normal_font
    ws_flags.cell(row=i, column=8, value=f['flags']).font = normal_font

    sev = f['severity']
    fill = critical_fill if sev == 'CRITICAL' else error_fill if sev == 'ERROR' else warning_fill if sev == 'WARNING' else gap_fill
    ws_flags.cell(row=i, column=1).fill = fill
    if sev == 'CRITICAL':
        ws_flags.cell(row=i, column=1).font = white_font

col_widths = [12, 14, 10, 14, 12, 8, 12, 80]
for j, w in enumerate(col_widths, 1):
    ws_flags.column_dimensions[chr(64+j)].width = w

# --- Sheet 3: All Opening Decisions (full audit trail) ---
ws_open = wb.create_sheet("Opening Ranges")

open_headers = ['Position', 'Stack (BB)', 'Hand', 'App Action', 'Flags']
for col, h in enumerate(open_headers, 1):
    cell = ws_open.cell(row=1, column=col, value=h)
    cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='333333')

for i, r in enumerate(opening_results, 2):
    ws_open.cell(row=i, column=1, value=r['position']).font = normal_font
    ws_open.cell(row=i, column=2, value=r['stack']).font = normal_font
    ws_open.cell(row=i, column=3, value=r['hand']).font = normal_font
    ws_open.cell(row=i, column=4, value=r['app_action']).font = normal_font
    ws_open.cell(row=i, column=5, value=r['flags']).font = normal_font
    if r['severity']:
        fill = critical_fill if r['severity'] == 'CRITICAL' else error_fill if r['severity'] == 'ERROR' else warning_fill
        for c in range(1, 6):
            ws_open.cell(row=i, column=c).fill = fill

for j, w in enumerate([12, 12, 8, 12, 60], 1):
    ws_open.column_dimensions[chr(64+j)].width = w

# --- Sheet 4: All Facing-Open Decisions ---
ws_face = wb.create_sheet("Facing Open Ranges")

face_headers = ['Opener', 'Hero', 'Stack (BB)', 'Hand', 'App Action', 'Flags']
for col, h in enumerate(face_headers, 1):
    cell = ws_face.cell(row=1, column=col, value=h)
    cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='333333')

for i, r in enumerate(facing_results, 2):
    ws_face.cell(row=i, column=1, value=r.get('opener', '')).font = normal_font
    ws_face.cell(row=i, column=2, value=r['position']).font = normal_font
    ws_face.cell(row=i, column=3, value=r['stack']).font = normal_font
    ws_face.cell(row=i, column=4, value=r['hand']).font = normal_font
    ws_face.cell(row=i, column=5, value=r['app_action']).font = normal_font
    ws_face.cell(row=i, column=6, value=r['flags']).font = normal_font
    if r['severity']:
        fill = critical_fill if r['severity'] == 'CRITICAL' else error_fill if r['severity'] == 'ERROR' else warning_fill if r['severity'] == 'WARNING' else gap_fill
        for c in range(1, 7):
            ws_face.cell(row=i, column=c).fill = fill

for j, w in enumerate([10, 10, 12, 8, 12, 80], 1):
    ws_face.column_dimensions[chr(64+j)].width = w

output_path = 'range-validation-report.xlsx'
wb.save(output_path)
print(f"\nReport saved to: {output_path}")
